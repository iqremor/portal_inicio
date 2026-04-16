# admin.py

# --Librerías Necesarias --
import os
import os.path as op
from datetime import datetime, timedelta

from flask import current_app, flash, jsonify, redirect, request, session, url_for
from flask_admin import Admin, AdminIndexView, BaseView, expose
from flask_admin.contrib.fileadmin import FileAdmin
from flask_admin.contrib.sqla import ModelView
from sqlalchemy.exc import OperationalError
from wtforms import PasswordField, TextAreaField

from models import (
    ActiveSession,
    Comentario,
    ConfiguracionSistema,
    Cuadernillo,
    ExamAnswer,
    ExamAvailability,
    ExamResult,
    Log,
    Peticion,
    PeticionEstado,
    User,
    UserRole,
    db,
)

# Se crea la instancia de SQLAlchemy sin asociarla a la app todavía
# (La instancia de Admin se crea y configura dentro de init_admin)


# Vista personalizada para el panel principal
class MyAdminIndexView(AdminIndexView):
    def is_accessible(self):
        if self.admin.app.debug:
            print(f"is_accessible called. logged_in: {session.get('logged_in')}")
        return session.get("logged_in")

    def inaccessible_callback(self, name, **kwargs):
        if self.admin.app.debug:
            print(
                f"inaccessible_callback called. endpoint: {request.endpoint}, url: {request.url}, path: {request.path}"
            )

            login_path = url_for("admin.login_view")
            print(f"login_path from url_for: {login_path}")

            if request.endpoint == "admin.login_view":
                print("Match: request.endpoint == 'admin.login_view'. Allowing login view to proceed.")

            print("No match: request.path != login_path. Redirecting to login with next parameter.")
        if request.endpoint == "admin.login_view":
            return None  # This line must be outside the debug block
        return redirect(url_for("admin.login_view", next=request.url))

    @expose("/")
    def index(self):
        # Obtener el estado del sitio de prueba
        test_site_enabled = ConfiguracionSistema.query.filter_by(clave="TEST_SITE_ENABLED").first()
        if not test_site_enabled:
            test_site_enabled = ConfiguracionSistema(
                clave="TEST_SITE_ENABLED",
                valor="0",
                descripcion="Activa o desactiva el sitio de prueba.",
            )
            db.session.add(test_site_enabled)
            db.session.commit()

        # Estadísticas básicas
        stats = {
            "total_usuarios": User.query.count(),
            "total_peticiones": Peticion.query.count(),
            "peticiones_pendientes": Peticion.query.filter_by(estado=PeticionEstado.PENDIENTE).count(),
            "peticiones_completadas": Peticion.query.filter_by(estado=PeticionEstado.COMPLETADA).count(),
        }

        # Peticiones recientes
        peticiones_recientes = Peticion.query.order_by(Peticion.created_at.desc()).limit(5).all()

        # Usuarios recientes
        usuarios_recientes = User.query.order_by(User.created_at.desc()).limit(5).all()

        return self.render(
            "admin/index.html",
            stats=stats,
            peticiones_recientes=peticiones_recientes,
            usuarios_recientes=usuarios_recientes,
            test_site_enabled=test_site_enabled.valor == "1",
        )

    @expose("/toggle-test-site")
    def toggle_test_site(self):
        config = ConfiguracionSistema.query.filter_by(clave="TEST_SITE_ENABLED").first()
        if config:
            config.valor = "1" if config.valor == "0" else "0"
            db.session.commit()
        return redirect(url_for(".index"))

    @expose("/login/", methods=("GET", "POST"))
    def login_view(self):
        # handle user login
        if request.method == "POST":
            username = request.form["username"]
            password = request.form["password"]

            user = db.session.query(User).filter_by(username=username, role=UserRole.ADMIN).first()

            if user and user.check_password(password):
                session["logged_in"] = True
                session["admin_user_id"] = user.id  # Store admin user ID in session
                session["user_role"] = user.role.value  # <-- AÑADIR ESTA LÍNEA
                if self.admin.app.debug:
                    print(
                        f"DEBUG: After setting session['logged_in'] in POST: {session.get('logged_in')}"
                    )  # Debug print
                    print(f"DEBUG: Full session object after login: {dict(session)}")  # Print full session
                flash("¡Inicio de sesión exitoso!", "success")
                return redirect(request.args.get("next") or url_for("admin.index"))
            else:
                flash("Nombre de usuario o contraseña incorrectos.", "danger")

        if self.admin.app.debug:
            print(f"DEBUG: session['logged_in'] on GET/initial load: {session.get('logged_in')}")  # Debug print
            print(f"DEBUG: Full session object on GET/initial load: {dict(session)}")  # Print full session
        return self.render("admin/login.html")

    @expose("/logout/")
    def logout_view(self):
        session.pop("logged_in", None)
        flash("Has cerrado sesión.", "info")
        return redirect(url_for("admin.login_view"))


# Vista personalizada para usuarios
class UserModelView(ModelView):
    # Configuración de columnas
    column_list = ["username", "role", "is_active", "created_at", "peticiones"]
    column_details_list = [
        "id",
        "username",
        "role",
        "is_active",
        "created_at",
        "updated_at",
    ]
    column_searchable_list = ["username"]
    column_filters = ["role", "is_active", "created_at"]
    column_editable_list = ["is_active", "role"]
    edit_template = "admin/user/edit.html"

    def is_accessible(self):
        """Verifica si el usuario tiene acceso al panel de administración"""
        from flask import session

        return session.get("logged_in") and session.get("user_role") == "admin"

    def inaccessible_callback(self, name, **kwargs):
        """Redirige a la página de login cuando no hay acceso"""
        return redirect(url_for("admin.login_view", next=request.url))

    # Configuración de formularios
    form_excluded_columns = [
        "password_hash",
        "peticiones",
        "comentarios",
        "logs",
        "updated_at",
    ]

    # Etiquetas personalizadas
    column_labels = {
        "username": "Nombre de Usuario",
        "role": "Rol",
        "is_active": "Activo",
        "created_at": "Fecha de Creación",
        "updated_at": "Última Actualización",
        "peticiones": "Peticiones",
    }

    # Formateo de columnas
    column_type_formatters = {datetime: lambda view, value: value.strftime("%d/%m/%Y %H:%M") if value else ""}

    # Configuración de paginación
    page_size = 25

    # Campo personalizado para contraseña
    form_extra_fields = {
        "password": PasswordField("Nueva Contraseña"),
    }

    def create_model(self, form):
        """Personaliza la creación de usuarios"""
        try:
            model = self.model()
            form.populate_obj(model)

            # Hashear la contraseña si se proporciona
            if form.password.data:
                model.set_password(form.password.data)

            self.session.add(model)
            self._on_model_change(form, model, True)
            self.session.commit()
        except Exception as ex:
            if not self.handle_view_exception(ex):
                raise
            self.session.rollback()
            return False
        else:
            self.after_model_change(form, model, True)
        return model

    def update_model(self, form, model):
        """Personaliza la actualización de usuarios"""
        try:
            form.populate_obj(model)

            # Actualizar contraseña solo si se proporciona una nueva
            if form.password.data:
                model.set_password(form.password.data)

            model.updated_at = datetime.utcnow()
            self._on_model_change(form, model, False)
            self.session.commit()
        except Exception as ex:
            if not self.handle_view_exception(ex):
                raise
            self.session.rollback()
            return False
        else:
            self.after_model_change(form, model, False)
        return True

    def edit_form(self, obj=None):
        form = super(UserModelView, self).edit_form(obj)
        from models import Cuadernillo, UserCuadernilloActivation

        # Obtener todos los cuadernillos
        todos_los_cuadernillos = Cuadernillo.query.all()

        # Obtener las activaciones para el usuario actual
        activaciones = UserCuadernilloActivation.query.filter_by(user_id=obj.id).all()
        activaciones_dict = {act.cuadernillo_id: act.is_active for act in activaciones}

        # Añadir información al formulario para que esté disponible en la plantilla
        form.cuadernillos = todos_los_cuadernillos
        form.activaciones = activaciones_dict

        return form


# Vista personalizada para peticiones
class PeticionModelView(ModelView):
    # Configuración de columnas
    column_list = [
        "titulo",
        "usuario",
        "estado",
        "prioridad",
        "created_at",
        "fecha_limite",
    ]
    column_details_list = [
        "id",
        "titulo",
        "descripcion",
        "usuario",
        "estado",
        "prioridad",
        "created_at",
        "updated_at",
        "fecha_limite",
    ]
    column_searchable_list = ["titulo", "descripcion"]
    column_filters = ["estado", "prioridad", "created_at", "user_id"]
    column_editable_list = ["estado", "prioridad"]

    def is_accessible(self):
        """Verifica si el usuario tiene acceso al panel de administración"""
        from flask import session

        return session.get("logged_in") and session.get("user_role") == "admin"

    def inaccessible_callback(self, name, **kwargs):
        """Redirige a la página de login cuando no hay acceso"""
        return redirect(url_for("admin.login_view", next=request.url))

    # Etiquetas personalizadas
    column_labels = {
        "titulo": "Título",
        "descripcion": "Descripción",
        "usuario": "Usuario",
        "estado": "Estado",
        "prioridad": "Prioridad",
        "created_at": "Fecha de Creación",
        "updated_at": "Última Actualización",
        "fecha_limite": "Fecha Límite",
        "user_id": "ID Usuario",
    }

    # Formateo de columnas
    column_type_formatters = {datetime: lambda view, value: value.strftime("%d/%m/%Y %H:%M") if value else ""}

    # Configuración de formularios
    form_overrides = {"descripcion": TextAreaField}

    # Configuración de paginación
    page_size = 25

    def get_query(self):
        """Personaliza la query base"""
        return self.session.query(self.model).join(User)

    def get_count_query(self):
        """Personaliza la query de conteo"""
        return self.session.query(self.model).join(User)


# Vista personalizada para comentarios
class ComentarioModelView(ModelView):
    column_list = ["peticion", "usuario", "contenido", "created_at"]
    column_details_list = ["id", "contenido", "peticion", "usuario", "created_at"]
    column_searchable_list = ["contenido"]
    column_filters = ["created_at", "user_id", "peticion_id"]

    column_labels = {
        "contenido": "Contenido",
        "peticion": "Petición",
        "usuario": "Usuario",
        "created_at": "Fecha de Creación",
    }

    def is_accessible(self):
        """Verifica si el usuario tiene acceso al panel de administración"""
        from flask import session

        return session.get("logged_in") and session.get("user_role") == "admin"

    def inaccessible_callback(self, name, **kwargs):
        """Redirige a la página de login cuando no hay acceso"""
        return redirect(url_for("admin.login_view", next=request.url))

    form_overrides = {"contenido": TextAreaField}

    column_type_formatters = {datetime: lambda view, value: value.strftime("%d/%m/%Y %H:%M") if value else ""}

    page_size = 25


# Vista para configuración del sistema
class ConfiguracionSistemaView(ModelView):
    """Vista personalizada para el modelo ConfiguracionSistema."""

    column_list = ("clave", "valor", "descripcion", "updated_at")
    form_columns = ("clave", "valor", "descripcion")
    column_labels = {
        "clave": "Clave",
        "valor": "Valor",
        "descripcion": "Descripción",
        "updated_at": "Última Modificación",
    }

    def is_accessible(self):
        """Verifica si el usuario tiene acceso al panel de administración"""
        from flask import session

        return session.get("logged_in") and session.get("user_role") == "admin"

    def inaccessible_callback(self, name, **kwargs):
        """Redirige a la página de login cuando no hay acceso"""
        return redirect(url_for("admin.login_view", next=request.url))


# Vista para Sesiones Activas
class ActiveSessionView(ModelView):
    list_template = "admin/active_session_list.html"
    # Deshabilitar la creación y edición manual desde el panel
    can_create = False
    can_edit = False
    can_delete = True  # Permitir eliminar para forzar el logout

    # Columnas a mostrar en la lista
    column_list = (
        "user.nombre_completo",
        "login_time",
        "last_seen",
        "ip_address",
        "user_agent",
        "cuadernillo.nombre",
    )  # Added cuadernillo.nombre

    # Etiquetas más amigables para las columnas
    column_labels = {
        "user.nombre_completo": "Usuario",
        "login_time": "Hora de Inicio",
        "last_seen": "Última Actividad",
        "ip_address": "Dirección IP",
        "user_agent": "Navegador/Dispositivo",
        "cuadernillo.nombre": "Examen Actual",  # Added label
    }

    # Ordenar por defecto por la última actividad
    column_default_sort = ("last_seen", True)

    # Formateo de fechas
    column_type_formatters = {
        datetime: lambda view, value, name: (
            (value - timedelta(hours=5)).strftime("%d/%m/%Y %H:%M:%S") if value else ""
        )
    }

    def is_accessible(self):
        """Verifica si el usuario tiene acceso."""
        return session.get("logged_in") and session.get("user_role") == "admin"

    def inaccessible_callback(self, name, **kwargs):
        """Redirige a la página de login cuando no hay acceso."""
        return redirect(url_for("web_main.login", next=request.url))


from flask import send_file

from utils.user_management import UserSyncManager

# ... (otras vistas)


class GestionUsuariosView(BaseView):
    @expose("/")
    def index(self):
        return self.render("admin/gestion_usuarios.html")

    @expose("/download_template")
    def download_template(self):
        try:
            template_path = os.path.join(current_app.instance_path, "plantilla_usuarios_iem.xlsx")
            UserSyncManager.generate_template(template_path)
            return send_file(template_path, as_attachment=True, download_name="plantilla_usuarios_iem.xlsx")
        except Exception as e:
            flash(f"Error al generar la plantilla: {str(e)}", "danger")
            return redirect(url_for(".index"))

    @expose("/import", methods=["POST"])
    def import_from_file(self):
        if "user_file" not in request.files:
            flash("No se seleccionó ningún archivo.", "danger")
            return redirect(url_for(".index"))

        file = request.files["user_file"]
        if file.filename == "":
            flash("Nombre de archivo vacío.", "danger")
            return redirect(url_for(".index"))

        temp_path = os.path.join(current_app.instance_path, file.filename)
        file.save(temp_path)

        delete_others = request.form.get("sync_mirror") == "on"

        try:
            summary = UserSyncManager.sync_from_file(temp_path, delete_others)
            session["last_focus_summary"] = summary
            flash("Sincronización masiva completada.", "success")
        except Exception as e:
            flash(f"Error durante la sincronización: {str(e)}", "danger")
        finally:
            if os.path.exists(temp_path):
                os.remove(temp_path)

        return redirect(url_for(".index"))

    @expose("/add_individual", methods=["POST"])
    def add_individual(self):
        codigo = request.form.get("codigo")
        nombre = request.form.get("nombre")
        grado = request.form.get("grado")

        if not codigo or not nombre or not grado:
            flash("Todos los campos son obligatorios.", "danger")
            return redirect(url_for(".index"))

        try:
            user = User.query.filter_by(codigo=codigo).first()
            if user:
                flash(f"El código {codigo} ya está registrado.", "warning")
            else:
                new_user = User(
                    username=codigo,
                    codigo=codigo,
                    nombre_completo=nombre,
                    grado=grado,
                    role=UserRole.USER,
                    is_active=True,
                )
                new_user.set_password(codigo)
                db.session.add(new_user)
                db.session.commit()
                flash(f"Usuario {nombre} añadido exitosamente.", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Error al añadir usuario: {str(e)}", "danger")

        return redirect(url_for(".index"))

    @expose("/clear_students", methods=["POST"])
    def clear_students(self):
        confirm = request.form.get("confirm_text", "").strip().upper()
        if confirm != "ELIMINAR":
            flash("La palabra de confirmación no coincide. No se realizaron cambios.", "warning")
            return redirect(url_for(".index"))

        try:
            count = UserSyncManager.clear_all_students()
            flash(f"Se han eliminado exitosamente {count} estudiantes de la base de datos.", "success")
        except Exception as e:
            flash(f"Error al limpiar la base de datos: {str(e)}", "danger")

        return redirect(url_for(".index"))

    def is_accessible(self):
        return session.get("logged_in") and session.get("user_role") == "admin"

    def inaccessible_callback(self, name, **kwargs):
        return redirect(url_for("admin.login_view", next=request.url))


class GestionIntentosView(BaseView):
    @expose("/")
    def index(self):
        return self.render("admin/gestion_intentos.html")

    def is_accessible(self):
        return session.get("logged_in") and session.get("user_role") == "admin"

    def inaccessible_callback(self, name, **kwargs):
        return redirect(url_for("admin.login_view", next=request.url))


class ConfigExamenesView(BaseView):
    @expose("/", methods=("GET", "POST"))
    def index(self):
        if request.method == "POST":
            # Extraer valor y convertir a ms
            try:
                sec_val = int(request.form.get("next_button_delay", 10))
                ms_val = sec_val * 1000
            except:
                ms_val = 10000

            configs = {
                "EXAM_TIMER_DURATION": request.form.get("timer_duration"),
                "EXAM_WARNING_TIME": request.form.get("warning_time"),
                "EXAM_NEXT_BUTTON_DELAY": str(ms_val),
                "EXAM_NUM_ATTEMPTS": request.form.get("num_attempts"),
                "EXAM_QUESTIONS_COUNT": request.form.get("questions_count"),
                "SHOW_CORRECT_ANSWERS": "1" if request.form.get("show_correct_answers") == "on" else "0",
                "SHOW_MARKED_ANSWERS": "1" if request.form.get("show_marked_answers") == "on" else "0",
                "PREICFES_ENABLED": "1" if request.form.get("preicfes_enabled") == "on" else "0",
                "PREUNAL_ENABLED": "1" if request.form.get("preunal_enabled") == "on" else "0",
                "LABORATORIOS_ENABLED": "1" if request.form.get("laboratorios_enabled") == "on" else "0",
            }

            try:
                for clave, valor in configs.items():
                    if valor is not None:
                        config = ConfiguracionSistema.query.filter_by(clave=clave).first()
                        if config:
                            config.valor = str(valor)
                            db.session.add(config)
                        else:
                            config = ConfiguracionSistema(
                                clave=clave, valor=str(valor), descripcion=f"Configuración de {clave}"
                            )
                            db.session.add(config)
                db.session.commit()
                flash("Configuración actualizada correctamente.", "success")
            except Exception as e:
                db.session.rollback()
                flash(f"Error al guardar: {str(e)}", "danger")

            return redirect(url_for("config_examenes.index"))

        # CARGA DE DATOS
        db.session.expire_all()

        def get_v(clave, default):
            c = ConfiguracionSistema.query.filter_by(clave=clave).first()
            return c.valor if c else default

        # Conversión de MS a Segundos para el panel
        try:
            db_ms = int(get_v("EXAM_NEXT_BUTTON_DELAY", "10000"))
            panel_sec = db_ms // 1000
        except:
            panel_sec = 10

        current_config = {
            "timer_duration": get_v("EXAM_TIMER_DURATION", "240"),
            "warning_time": get_v("EXAM_WARNING_TIME", "30"),
            "next_button_delay": panel_sec,
            "num_attempts": get_v("EXAM_NUM_ATTEMPTS", "1"),
            "questions_count": get_v("EXAM_QUESTIONS_COUNT", "10"),
            "show_correct_answers": get_v("SHOW_CORRECT_ANSWERS", "0") == "1",
            "show_marked_answers": get_v("SHOW_MARKED_ANSWERS", "0") == "1",
            "preicfes_enabled": get_v("PREICFES_ENABLED", "0") == "1",
            "preunal_enabled": get_v("PREUNAL_ENABLED", "0") == "1",
            "laboratorios_enabled": get_v("LABORATORIOS_ENABLED", "0") == "1",
        }
        return self.render("admin/config_examenes.html", exam_settings=current_config)

    def is_accessible(self):
        return session.get("logged_in") and session.get("user_role") == "admin"

    def inaccessible_callback(self, name, **kwargs):
        return redirect(url_for("admin.login_view", next=request.url))


class ReporteGradoView(BaseView):
    @expose("/")
    def index(self):
        # Obtener todos los grados únicos de la tabla User
        grados_query = db.session.query(User.grado).filter(User.grado != None).distinct().all()
        grados = sorted([g[0] for g in grados_query if g[0]])

        # Obtener todos los cuadernillos (para el dropdown inicial, cargamos todos)
        cuadernillos_objs = Cuadernillo.query.order_by(Cuadernillo.grado, Cuadernillo.area).all()
        cuadernillos_list = [c.to_dict() for c in cuadernillos_objs]

        return self.render("admin/reporte_grado.html", grados=grados, cuadernillos=cuadernillos_list)

    @expose("/api/examenes_por_grado/<string:grado>")
    def api_examenes_por_grado(self, grado):
        """API interna para cargar el dropdown de exámenes asignados a un grado."""
        availability = ExamAvailability.query.filter_by(grado=grado, is_enabled=True).all()
        data = []
        for a in availability:
            if a.cuadernillo:
                d = a.cuadernillo.to_dict()
                # Añadir etiqueta de origen para la UI
                d["label"] = f"{d['area']} (Origen: Saber {d['grado']}) - {d['nombre']}"
                data.append(d)
        return jsonify(data)

    @expose("/exportar_excel/<string:grado>/<int:cuadernillo_id>")
    def exportar_excel(self, grado, cuadernillo_id):
        from datetime import datetime
        from io import BytesIO

        from flask import send_file
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        # 1. Obtener datos
        usuarios = User.query.filter_by(grado=grado).all()
        cuadernillo = Cuadernillo.query.get(cuadernillo_id)

        if not cuadernillo:
            return "Cuadernillo no encontrado", 404

        wb = Workbook()
        ws = wb.active
        ws.title = f"Reporte {grado}"

        # Estilos
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center")
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Encabezados de información general
        ws.merge_cells("A1:F1")
        ws["A1"] = f"REPORTE DE RESULTADOS - GRADO {grado}"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].alignment = center_align

        ws.merge_cells("A2:F2")
        ws["A2"] = f"Examen: {cuadernillo.area.upper()} (Origen: Grado {cuadernillo.grado}) - {cuadernillo.nombre}"
        ws["A2"].font = Font(size=12)
        ws["A2"].alignment = center_align

        ws.merge_cells("A3:F3")
        ws["A3"] = f"Fecha de generación: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A3"].alignment = center_align

        # Encabezados de tabla
        headers = ["Código", "Nombre Completo", "Nota Final", "Porcentaje", "Intentos", "Fecha Finalización"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        # Datos
        row_num = 6
        notas = []
        for usuario in usuarios:
            ultimo_result = (
                ExamResult.query.filter_by(user_id=usuario.id, cuadernillo_id=cuadernillo_id)
                .order_by(ExamResult.attempt_number.desc())
                .first()
            )

            intentos_count = ExamResult.query.filter_by(user_id=usuario.id, cuadernillo_id=cuadernillo_id).count()

            nota = ultimo_result.final_score if ultimo_result else None
            if nota is not None:
                notas.append(nota)

            ws.cell(row=row_num, column=1, value=usuario.codigo).border = border
            ws.cell(row=row_num, column=2, value=usuario.nombre_completo).border = border

            nota_cell = ws.cell(row=row_num, column=3, value=round(nota, 1) if nota is not None else "N/A")
            nota_cell.border = border
            nota_cell.alignment = center_align

            if nota is not None:
                if nota >= 3.0:
                    nota_cell.fill = green_fill
                else:
                    nota_cell.fill = red_fill

                porcentaje = (nota / 5.0) * 100
                ws.cell(row=row_num, column=4, value=f"{round(porcentaje, 1)}%").border = border
            else:
                ws.cell(row=row_num, column=4, value="-").border = border

            ws.cell(row=row_num, column=4).alignment = center_align

            ws.cell(row=row_num, column=5, value=intentos_count).border = border
            ws.cell(row=row_num, column=5).alignment = center_align

            fecha_str = ultimo_result.completion_date.strftime("%Y-%m-%d %H:%M") if ultimo_result else "-"
            ws.cell(row=row_num, column=6, value=fecha_str).border = border
            ws.cell(row=row_num, column=6).alignment = center_align

            row_num += 1

        # Resumen / Promedio
        row_num += 1
        ws.cell(row=row_num, column=2, value="PROMEDIO GRUPAL").font = Font(bold=True)
        promedio = sum(notas) / len(notas) if notas else 0
        promedio_cell = ws.cell(row=row_num, column=3, value=round(promedio, 2))
        promedio_cell.font = Font(bold=True)
        promedio_cell.alignment = center_align
        if promedio >= 3.0:
            promedio_cell.fill = green_fill
        else:
            promedio_cell.fill = red_fill

        # Ajustar ancho de columnas
        column_widths = [15, 45, 12, 12, 10, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=5, column=i).column_letter].width = width

        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Reporte_UNICUS_{grado}_{cuadernillo.area}_{datetime.now().strftime('%Y%m%d')}.xlsx"

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )

    @expose("/api/resultados/<string:grado>/<int:cuadernillo_id>")
    def get_resultados_json(self, grado, cuadernillo_id):
        """Devuelve los resultados de un grado y examen en formato JSON."""
        if not self.is_accessible():
            return jsonify({"error": "No autorizado"}), 403

        try:
            usuarios = User.query.filter_by(grado=grado).all()
            resultados_data = []

            for usuario in usuarios:
                ultimo_result = (
                    ExamResult.query.filter_by(user_id=usuario.id, cuadernillo_id=cuadernillo_id)
                    .order_by(ExamResult.attempt_number.desc())
                    .first()
                )

                intentos_count = ExamResult.query.filter_by(user_id=usuario.id, cuadernillo_id=cuadernillo_id).count()

                resultados_data.append(
                    {
                        "codigo": usuario.codigo,
                        "nombre": usuario.nombre_completo,
                        "nota": ultimo_result.final_score if ultimo_result else None,
                        "intentos": intentos_count,
                        "fecha": ultimo_result.completion_date.strftime("%Y-%m-%d %H:%M") if ultimo_result else None,
                    }
                )

            return jsonify({"resultados": resultados_data})
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    @expose("/limpiar_notas/<string:grado>/<int:cuadernillo_id>", methods=["POST"])
    def limpiar_notas(self, grado, cuadernillo_id):
        """Elimina todos los resultados y respuestas para un grado y cuadernillo específicos."""
        try:
            # 1. Obtener usuarios del grado
            usuarios = User.query.filter_by(grado=grado).all()
            user_ids = [u.id for u in usuarios]

            if not user_ids:
                flash(f"No se encontraron usuarios para el grado {grado}.", "warning")
                return redirect(url_for(".index"))

            # 2. Eliminar Resultados
            results_to_delete = ExamResult.query.filter(
                ExamResult.user_id.in_(user_ids), ExamResult.cuadernillo_id == cuadernillo_id
            ).all()

            for result in results_to_delete:
                # 3. Eliminar Respuestas asociadas (por ID de sesión si existe relación,
                # o por user_id y cuadernillo_id)
                ExamAnswer.query.filter_by(user_id=result.user_id, cuadernillo_id=result.cuadernillo_id).delete()
                db.session.delete(result)

            # 4. Resetear Sesiones Activas (importante para permitir reintento)
            ActiveSession.query.filter(
                ActiveSession.user_id.in_(user_ids), ActiveSession.cuadernillo_id == cuadernillo_id
            ).delete(synchronize_session=False)

            db.session.commit()
            flash(
                f"Se han eliminado exitosamente los resultados del grado {grado} para este examen.",
                "success",
            )
        except Exception as e:
            db.session.rollback()
            flash(f"Error al limpiar las notas: {str(e)}", "danger")

        return redirect(url_for(".index"))

    def is_accessible(self):
        return session.get("logged_in") and session.get("user_role") == "admin"

    def inaccessible_callback(self, name, **kwargs):
        return redirect(url_for("admin.login_view", next=request.url))


def init_admin(app):
    """Inicializa Flask-Admin."""
    admin = Admin(
        app,
        name="Panel de Administración",
        template_mode="bootstrap4",
        index_view=MyAdminIndexView(url="/admin"),
    )

    # Registrar las vistas de modelos
    admin.add_view(UserModelView(User, db.session, name="Usuarios"))
    admin.add_view(ModelView(Cuadernillo, db.session, name="Cuadernillos", endpoint="cuadernillo"))
    admin.add_view(PeticionModelView(Peticion, db.session, name="Peticiones"))
    admin.add_view(ComentarioModelView(Comentario, db.session, name="Comentarios"))
    admin.add_view(ConfiguracionSistemaView(ConfiguracionSistema, db.session, name="Configuración Sistema"))
    admin.add_view(GestionUsuariosView(name="Gestión de Usuarios (FOCUS)", endpoint="gestion_usuarios"))
    admin.add_view(ConfigExamenesView(name="Configuración de Exámenes", endpoint="config_examenes"))
    admin.add_view(ReporteGradoView(name="Reporte por Grado", endpoint="reporte_grado"))
    admin.add_view(ModelView(Log, db.session, name="Logs del Sistema"))
    admin.add_view(ActiveSessionView(ActiveSession, db.session, name="Sesiones Activas"))
    admin.add_view(DatabaseAdminView(name="Gestión DB", endpoint="db_admin"))
    admin.add_view(ExamAvailabilityView(name="Gestión de Exámenes", endpoint="exam_availability"))
    admin.add_view(GestionIntentosView(name="Gestión de Intentos", endpoint="gestion_intentos"))

    # Añadir vista de gestión de archivos para la raíz del proyecto
    path = op.abspath(op.join(op.dirname(__file__), ".."))
    admin.add_view(FileAdmin(path, "/files/", name="Gestor de Archivos"))

    return admin


# Agregar en la función init_admin o donde configures las vistas


class AdminDashboardView(BaseView):
    @expose("/")
    def index(self):
        # Agregar lógica para mostrar estado
        return self.render("admin/index.html")


# La plantilla del dashboard ahora se carga desde templates/admin/index.html
# y ya no es necesario generarla desde aquí. El código anterior ha sido eliminado.


class DatabaseAdminView(BaseView):
    @expose("/", methods=("GET", "POST"))
    def index(self):
        if request.method == "POST":
            # Manejar carga de archivos
            if "database" in request.files:
                file = request.files["database"]
                if file.filename != "":
                    if file.filename.endswith(".db"):
                        import os

                        from flask import current_app

                        file_path = os.path.join(current_app.instance_path, "sistema_gestion.db")
                        file.save(file_path)
                        flash(
                            "Base de datos cargada y reemplazada exitosamente.",
                            "success",
                        )
                    else:
                        flash("Error: El archivo debe tener la extensión .db", "danger")
                else:
                    flash("No se seleccionó ningún archivo para cargar.", "warning")
                return redirect(url_for(".index"))

            action = request.form.get("action")
            try:
                if action == "apply_prefix":
                    prefix = request.form.get("path_prefix")
                    if not prefix:
                        flash("No se proporcionó ningún prefijo.", "warning")
                    else:
                        cuadernillos = Cuadernillo.query.all()

                        actualizados = 0
                        for c in cuadernillos:
                            if not c.dir_banco.startswith(prefix):
                                c.dir_banco = f"{prefix}{c.dir_banco}"
                                actualizados += 1

                        if actualizados > 0:
                            db.session.commit()
                            flash(
                                f'Se ha aplicado el prefijo "{prefix}" a {actualizados} cuadernillo(s).',
                                "success",
                            )
                        else:
                            flash(
                                f'No se necesitó actualizar ningún cuadernillo con el prefijo "{prefix}".',
                                "info",
                            )

                elif action == "seed_db":
                    from models import seed_data

                    seed_data()
                    flash(
                        "La base de datos ha sido poblada con datos iniciales (seeding).",
                        "success",
                    )

                elif action == "create_tables":
                    from models import create_tables

                    create_tables()
                    flash(
                        "Todas las tablas han sido creadas en la base de datos.",
                        "success",
                    )

                elif action == "drop_tables":
                    from models import drop_tables

                    drop_tables()
                    flash(
                        "¡ADVERTENCIA! Todas las tablas han sido eliminadas de la base de datos.",
                        "danger",
                    )

                else:
                    flash("Acción desconocida.", "danger")

            except Exception as e:
                flash(f"Ocurrió un error: {str(e)}", "danger")

            return redirect(url_for(".index"))

        return self.render("admin/db_admin.html")

    def is_accessible(self):
        return session.get("logged_in") and session.get("user_role") == "admin"

    def inaccessible_callback(self, name, **kwargs):
        return redirect(url_for("admin.login_view", next=request.url))

    @expose("/download-db/")
    def download_db(self):
        from flask import current_app, send_from_directory

        db_path = current_app.instance_path
        return send_from_directory(db_path, "sistema_gestion.db", as_attachment=True)


class ExamAvailabilityView(BaseView):
    @expose("/", methods=("GET", "POST"))
    def index(self):
        if request.method == "POST":
            # 1. Obtener todos los grados únicos para procesar la matriz
            grados_query = db.session.query(User.grado).filter(User.grado != None).distinct().all()
            unique_grados = sorted([g[0] for g in grados_query if g[0]])

            # 2. Manejar Cuadernillos (Matriz Universal)
            all_cuadernillos = Cuadernillo.query.all()

            # Limpiar asignaciones anteriores si se desea sincronización espejo,
            # o actualizar individualmente. Para UNICUS usaremos actualización inteligente.
            for c in all_cuadernillos:
                for g in unique_grados:
                    # El campo en el form será: cuadernillo-{id}-grado-{g}
                    field_name = f"cuadernillo-{c.id}-grado-{g}"
                    is_enabled = field_name in request.form

                    availability = ExamAvailability.query.filter_by(cuadernillo_id=c.id, grado=g).first()

                    if is_enabled:
                        if not availability:
                            availability = ExamAvailability(cuadernillo_id=c.id, grado=g, is_enabled=True)
                            db.session.add(availability)
                        else:
                            availability.is_enabled = True
                    else:
                        if availability:
                            # En lugar de borrar, marcamos como deshabilitado para mantener histórico si es necesario,
                            # o borramos si preferimos limpiar la DB. Borraremos para limpieza.
                            db.session.delete(availability)

            # 3. Manejar Módulos Globales (Preicfes, Preunal, Laboratorios)
            modulos = ["PREICFES", "PREUNAL", "LABORATORIOS"]
            for mod in modulos:
                enabled_grades = []
                for g in unique_grados:
                    if f"module-{mod}-{g}" in request.form:
                        enabled_grades.append(str(g))

                clave = f"MODULE_{mod}_GRADES"
                valor = ",".join(enabled_grades)
                config = ConfiguracionSistema.query.filter_by(clave=clave).first()
                if config:
                    config.valor = valor
                else:
                    config = ConfiguracionSistema(
                        clave=clave, valor=valor, descripcion=f"Grados habilitados para el módulo {mod}"
                    )
                db.session.add(config)

            try:
                db.session.commit()
                flash("La matriz de asignación universal ha sido actualizada.", "success")
            except Exception as e:
                db.session.rollback()
                flash(f"Error al actualizar la disponibilidad: {str(e)}", "danger")

            active_filter = request.form.get("active_grade_filter", "all")
            return redirect(url_for(".index", filter_grade=active_filter))

        # Lógica GET
        try:
            availability_data = ExamAvailability.query.all()
        except OperationalError:
            flash("Error al acceder a la tabla de disponibilidad.", "warning")
            return self.render("admin/exam_availability.html", cuadernillos=[], grados=[], availability_map={})

        # Cargar todos los cuadernillos (los 27)
        cuadernillos_all = Cuadernillo.query.order_by(Cuadernillo.grado, Cuadernillo.area).all()

        # Categorizar para las pestañas
        cuadernillos = {
            "SABER": [c for c in cuadernillos_all if "pruebas_saber" in c.cuadernillo_id.lower()],
            "UNAL": [c for c in cuadernillos_all if "pruebas_unal" in c.cuadernillo_id.lower()],
            "LAB": [
                c
                for c in cuadernillos_all
                if "laboratorios" in c.cuadernillo_id.lower() or "lab" in c.cuadernillo_id.lower()
            ],
        }

        # Obtener todos los grados posibles
        grados_query = db.session.query(User.grado).filter(User.grado != None).distinct().all()
        grados = sorted([g[0] for g in grados_query if g[0]])

        # Mapa de disponibilidad: (cuadernillo_id, grado_destino) -> True/False
        availability_map = {}
        for avail in availability_data:
            availability_map[(avail.cuadernillo_id, avail.grado)] = avail.is_enabled

        # Cargar configuración de módulos
        def get_mod_grades(mod):
            c = ConfiguracionSistema.query.filter_by(clave=f"MODULE_{mod}_GRADES").first()
            return [g.strip() for g in c.valor.split(",")] if c and c.valor else []

        modules_config = {
            "PREICFES": get_mod_grades("PREICFES"),
            "PREUNAL": get_mod_grades("PREUNAL"),
            "LABORATORIOS": get_mod_grades("LABORATORIOS"),
        }

        return self.render(
            "admin/exam_availability.html",
            cuadernillos=cuadernillos,
            grados=grados,
            availability_map=availability_map,
            modules_config=modules_config,
        )

    def is_accessible(self):
        return session.get("logged_in") and session.get("user_role") == "admin"

    def inaccessible_callback(self, name, **kwargs):
        return redirect(url_for("admin.login_view", next=request.url))
